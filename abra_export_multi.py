from tkinter import Tk, Frame, Button, filedialog
from openpyxl import load_workbook


class Data:

	CATEGORY_COLUMNS = 0
	PARAMETERS_ROWS = 1
	PARAMETERS_VALUES_COLUMNS = 2
	PRODUCT_ROWS = 3

	def __init__(self):
		# self.user_selected_file = load_workbook("TV držáky.xlsx", data_only=True)
		self._window = Tk()
		self._window.resizable(False, False)
		self._window.eval("tk::PlaceWindow . center")
		self._window.title("Generátor importních CSV")
		self._main_frame = Frame(master=self._window, height=100, width=250).grid(sticky="nwse")
		self._select_button = Button(master=self._main_frame, text="Výběr souboru",
									 command=self.load_file).grid(row=0, column=0)
		self._window.mainloop()

	def load_file(self):
		file = filedialog.askopenfilename(filetypes=[("Excel soubory", "*.xlsx")])
		self._window.quit()
		self.user_selected_file = load_workbook(file, data_only=True)
		return self.user_selected_file

	def settings(self):
		settings_file = "settings.txt"
		with open(settings_file, "r", encoding="utf-8") as file:
			values = []
			reader = file.readlines()
			for line in reader:
				numeric_value = int(line.split(':')[1].replace(" ", "").replace("\n", ""))
				values.append(numeric_value)
			for line in reader[4:]:
				values.append(line.split(': ')[1])
			categories_columns = values[0]
			parameters_rows = values[1]
			parameters_values_columns = values[2]
			rows_for_products = values[3]
			return categories_columns, parameters_rows, parameters_values_columns, rows_for_products

	def get_data_from_cells(self, sheet_index, cols_or_rows, min_row, max_row, min_col, max_col, file, none_values=False):
		sheet_index = sheet_index
		cols_or_rows = cols_or_rows
		min_row = min_row
		max_row = max_row
		min_col = min_col
		max_col = max_col

		wb = file
		wb.active = sheet_index  # 0 = first sheet, 1 = second sheet
		ws = wb.active

		if cols_or_rows == "rows":
			for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col,
									max_col=max_col, values_only=True):
				if min_row == 1:
					for i, value in enumerate(row):
						if value is None:
							row = list(row)
							row[i] = ""
					return row
				if none_values is True:
					return row
				else:
					return [cell for cell in row if cell is not None]
		elif cols_or_rows == "cols":
			for col in ws.iter_cols(min_row=min_row, max_row=max_row, min_col=min_col,
									max_col=max_col, values_only=True):
				if none_values is True:
					return col
				else:
					return [cell for cell in col if cell is not None]
		else:
			return None


	def create_file_direct_values_of_cards(self):
		output_file = "parametry.csv"
		cz_parameters = self.get_data_from_cells(0, "rows", 4, 4, 3, 300, file=self.user_selected_file)[::3]
		sk_parameters = self.get_data_from_cells(0, "rows", 4, 4, 3, 300, file=self.user_selected_file)[1::3]
		en_parameters = self.get_data_from_cells(0, "rows", 4, 4, 3, 300, file=self.user_selected_file)[2::3]
		products = self.get_data_from_cells(0, "cols", 5, 300, 1, 1, file=self.user_selected_file)
		parameters_types = self.get_data_from_cells(0, "rows", 2, 2, 3, 300, file=self.user_selected_file)[::3]
		parameters_units = self.get_data_from_cells(0, "rows", 1, 1, 3, 300, file=self.user_selected_file)[::3]
		product_type = self.get_data_from_cells(0, "rows", 5, 5, 2, 2, file=self.user_selected_file)
		with open(output_file, mode="w", encoding="ansi") as file:
			file.writelines("Skladová karta;Vlastnost CZ;Vlastnost SK;Vlastnost EN;CZ Hodnota z číselníku;SK Hodnota z číselníku;EN Hodnota z číselníku;Hodnota;Jednotka\n")
		for product in range(len(products)):
			cz_dirty_valid_values = list(self.get_data_from_cells(0, 'rows', product+5, product+5, 0, 0, self.user_selected_file, none_values=True)[2::3])
			for par_ind in range(len(cz_parameters)):
				if cz_dirty_valid_values[par_ind] is None:
					cz_dirty_valid_values[par_ind] = ""
			cz_valid_values = [v for v in cz_dirty_valid_values if v is not None]
			sk_dirty_valid_values = list(self.get_data_from_cells(0, 'rows', product+5, product+5, 0, 0, self.user_selected_file, none_values=True)[3::3])
			for par_ind in range(len(sk_parameters)):
				if sk_dirty_valid_values[par_ind] is None:
					sk_dirty_valid_values[par_ind] = ""
			sk_valid_values = [v for v in sk_dirty_valid_values if v is not None]
			en_dirty_valid_values = list(self.get_data_from_cells(0, 'rows', product+5, product+5, 0, 0, self.user_selected_file, none_values=True)[4::3])
			for par_ind in range(len(en_parameters)):
				if en_dirty_valid_values[par_ind] is None:
					en_dirty_valid_values[par_ind] = ""
			en_valid_values = [v for v in en_dirty_valid_values if v is not None]
			for valid_value in range(len(cz_valid_values)):
				with open(output_file, mode="a", encoding="ansi") as file:
					if parameters_types[valid_value] == "Multihodnota":
						cz_multi_values = list(cz_valid_values[valid_value].split("|"))
						sk_multi_values = list(sk_valid_values[valid_value].split("|"))
						en_multi_values = list(en_valid_values[valid_value].split("|"))
						for multi_value in zip(cz_multi_values, sk_multi_values, en_multi_values):
							file.writelines(f"{products[product]};{cz_parameters[valid_value]};{sk_parameters[valid_value]};{en_parameters[valid_value]};{multi_value[0]};{multi_value[1]};{multi_value[2]};;{parameters_units[valid_value]}\n")
					elif parameters_types[valid_value] == "Číselníková hodnota" or parameters_types[valid_value] == "Ano/Ne":
						file.writelines(f"{products[product]};{cz_parameters[valid_value]};{sk_parameters[valid_value]};{en_parameters[valid_value]};{cz_valid_values[valid_value]};{sk_valid_values[valid_value]};{en_valid_values[valid_value]};;{parameters_units[valid_value]}\n")
					else:
						file.writelines(f"{products[product]};{cz_parameters[valid_value]};{sk_parameters[valid_value]};{en_parameters[valid_value]};;;;{cz_valid_values[valid_value]};{parameters_units[valid_value]}\n")


w = Data()
w.create_file_direct_values_of_cards()

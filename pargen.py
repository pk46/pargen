#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os


class Pargen:
    INDEX_CZ = 0
    INDEX_SK = 2
    INDEX_EN = 3
    LANGUAGE_INDEXES = (INDEX_CZ, INDEX_SK, INDEX_EN)
    LANGUAGE_NAMES = ("CZ", "SK", "EN")
    DV_WARNING_TEXT = "Tato hodnota neodpovídá omezením, která jsou pro tuto buňku nadefinovaná. Multihodnota se musí" \
                      " skládat z hodnot, které odpovídají omezením a musejí být odděleny svislítkem (alt+124) bez mezer."
    FILE_OVERWRITE_TEXT = ' existuje. Chceš jej přepsat? a/n '

    def __init__(self):
        self._settings = self.settings()
        # self._excel_file = ""
        self._wb = load_workbook(self._excel_file, data_only=True)
        self._categories = self.get_categories()
        self._choice = self.select_product_category()
        self._category_indicies = self.get_indicies_of_parameters()
        self._category_parameters = self.get_category_parameters()
        self._first_row_unit = ["Jednotka", None]
        self._second_row_type = ["Typ parametru", None]
        self._third_row_language = ["Jazyk", None]
        self._fourth_row_parameter = ["Kód produktu", "Typ produktu"]
        self._fifth_row_menu_data = [None]
        self._wb_output = Workbook()
        self._ws1 = self._wb_output.active

    def settings(self):
        settings_file = "settings.txt"
        with open(settings_file, "r", encoding="utf-8") as file:
            values = []
            reader = file.readlines()
            for line in reader[:4]:
                numeric_value = int(line.split(':')[1].replace(" ", "").replace("\n", ""))
                values.append(numeric_value)
            for line in reader[4:]:
                values.append(line.split(': ')[1])
            categories_columns = values[0]
            parameters_rows = values[1]
            parameters_values_columns = values[2]
            rows_for_products = values[3]
            self._excel_file = values[4]
            return categories_columns, parameters_rows, parameters_values_columns, rows_for_products

    def get_data_from_cells(self, sheet_index, cols_or_rows, min_row, max_row, min_col, max_col):
        _sheet_index = sheet_index
        _cols_or_rows = cols_or_rows
        _min_row = min_row
        _max_row = max_row
        _min_col = min_col
        _max_col = max_col

        self._wb.active = _sheet_index  # 0 = first sheet, 1 = second sheet
        _ws = self._wb.active

        if _cols_or_rows == "rows":
            for row in _ws.iter_rows(min_row=_min_row, max_row=_max_row, min_col=_min_col,
                                     max_col=_max_col, values_only=True):
                return row
        elif _cols_or_rows == "cols":
            for col in _ws.iter_cols(min_row=_min_row, max_row=_max_row, min_col=_min_col,
                                     max_col=_max_col, values_only=True):
                return col
        else:
            return None

    def get_categories(self):
        _categories = []
        _dirty_data = self.get_data_from_cells(1, "rows", 1, 1, 6, self._settings[0])

        for value in _dirty_data:
            if value is not None:
                _categories.append(value)
        return _categories  # returns cz product category names

    def get_all_parameters(self):
        _parameters = []
        _dirty_data_cz = self.get_data_from_cells(1, "cols", 4, self._settings[1], 3, 5)
        _dirty_data_sk = self.get_data_from_cells(1, "cols", 4, self._settings[1], 4, 5)
        _dirty_data_en = self.get_data_from_cells(1, "cols", 4, self._settings[1], 5, 5)

        for value in zip(_dirty_data_cz, _dirty_data_sk, _dirty_data_en):
            if value is not None:
                _parameters.append(value)
        return _parameters  # returns CZ, SK and EN names of all parameters

    def get_indicies_of_parameters(self):
        _category = self._choice
        _category_index = self._categories.index(_category) + 6  # +6 to skip unecessary columns
        _category_list_of_a_cells = self.get_data_from_cells(1, "cols", 4, self._settings[1],
                                                             _category_index,
                                                             _category_index)  # "A" cells for certain category
        _indicies_of_a_cells = [i for i, x in enumerate(_category_list_of_a_cells) if
                                x == "A"]  # indicies of A cells
        return _indicies_of_a_cells

    def get_category_parameters(self):
        _parameters = []
        _category_indicies = self._category_indicies
        for i in _category_indicies:
            _parameters.append(self.get_all_parameters()[i])
        return _parameters  # returns CZ, SK and EN parameters names for a given product category

    def get_parameters_type(self):
        _valid_types = []
        _all_types = self.get_data_from_cells(1, "cols", 4, self._settings[1], 1, 1)
        _category_indicies = self._category_indicies
        for i in _category_indicies:
            _valid_types.append(_all_types[i])
        return _valid_types  # returns types of all parameters of a given category

    def get_units(self):
        _valid_units = []
        _all_types = self.get_data_from_cells(1, "cols", 4, self._settings[1], 2, 2)
        _category_indicies = self._category_indicies
        for i in _category_indicies:
            _valid_units.append(_all_types[i])
        return _valid_units  # returns units of numeric parameters of a given category

    def get_valid_values_of_text_parameter_type(self):
        _category_choice = self._choice
        _product_types = self.get_data_from_cells(0, "cols", 2, self._settings[1], 5, 5)
        _category_indicies = [i + 2 for i, x in enumerate(_product_types) if x == _category_choice]  # +2 to skip rows
        valid_options = {}
        for row in _category_indicies:
            valid_options = self.get_parameter_options(valid_options, row)
        return valid_options

    def get_parameter_options(self, options, row):
        for language_index in Pargen.LANGUAGE_INDEXES:
            if language_index not in options:
                options[language_index] = {}
            name = self.get_data_from_cells(language_index, "rows", row, row, 2, 2)
            all_options = self.get_data_from_cells(language_index, "rows", row, row, 6, self._settings[2])
            valid_options = [value for value in all_options if value is not None]
            options[language_index][name] = valid_options
        return options

    def select_product_category(self):
        _category = self._categories
        for i, category in enumerate(_category):
            print(f"{i + 1}. {category}")
        while True:
            try:
                choice = int(input("\nVyber číslo kategorie \n"))
                try:
                    if choice not in range(1, len(_category) + 1):
                        print("Prosím zadej platné číslo kategorie ")
                        continue
                    break
                except ValueError:
                    pass
            except ValueError:
                print("Prosím zadej číselnou hodnotu kategorie")
        return _category[choice - 1]

    def prepare_data(self):
        for unit in self.get_units():
            for i in range(len(Pargen.LANGUAGE_NAMES)):
                self._first_row_unit.append(unit)

        for par_type in self.get_parameters_type():
            for i in range(len(Pargen.LANGUAGE_NAMES)):
                self._second_row_type.append(par_type)

        for i in range(len(self._category_parameters)):
            for lang in Pargen.LANGUAGE_NAMES:
                self._third_row_language.append(lang)

        for par_tuple in self._category_parameters:
            for parameter in par_tuple:
                self._fourth_row_parameter.append(parameter)

        self._fifth_row_menu_data.append(self._choice)

    def get_columns_to_write_data(self, language):
        text_indicies = []
        columns = []
        for i, types in enumerate(self._second_row_type):
            if types == "Číselníková hodnota" or types == "Multihodnota":
                if self._third_row_language[i] == language:
                    text_indicies.append(i)
        for ind in text_indicies:
            temp = []
            for i in range(5, self._settings[3] + 1):
                temp.append(f"{get_column_letter(ind + 1)}{i}")
            columns.append(temp)
        return columns

    def create_text_parameters_dropdowns(self):
        valid_parameters = self.get_valid_values_of_text_parameter_type()
        language_counter = 0
        for language in Pargen.LANGUAGE_INDEXES:
            category_counter = 0
            for category_index, category_in_all_languages in enumerate(self._category_parameters):
                category = (category_in_all_languages[language_counter], )
                try:
                    if category in valid_parameters[language].keys():
                        parameters = valid_parameters[language][category]
                        for i, value in enumerate(parameters):
                            if value is not str:
                                parameters[i] = str(value)
                        values = '"' + ','.join(parameters) + '"'
                        reference_matrix = self.get_columns_to_write_data(Pargen.LANGUAGE_NAMES[language_counter])
                        if self.get_parameters_type()[category_index] == "Multihodnota":
                            data_values = DataValidation(sqref=reference_matrix[category_counter], type="list", formula1=values, errorStyle="warning", error=Pargen.DV_WARNING_TEXT)
                            category_counter += 1
                            self._ws1.add_data_validation(data_values)
                        elif self.get_parameters_type()[category_index] == "Číselníková hodnota":
                            data_values = DataValidation(sqref=reference_matrix[category_counter], type="list", formula1=values)
                            category_counter += 1
                            self._ws1.add_data_validation(data_values)
                except KeyError:
                    pass
            language_counter += 1

    def create_yes_no_dropdowns(self, language):
        indicies = []
        columns = []
        yes_no_values = ["\"Ano, Ne\"", "\"Áno, Nie\"", "\"Yes, No\""]
        for i, types in enumerate(self._second_row_type):
            if types == "Ano/Ne":
                if self._third_row_language[i] == language:
                    indicies.append(i)
        for ind in indicies:
            temp = []
            for i in range(5, self._settings[3] + 1):
                temp.append(f"{get_column_letter(ind + 1)}{i}")
            columns.append(temp)
        if language == "CZ":
            for column in columns:
                data_vals = DataValidation(sqref=column, type="list", formula1=yes_no_values[0])
                self._ws1.add_data_validation(data_vals)
        elif language == "SK":
            for column in columns:
                data_vals = DataValidation(sqref=column, type="list", formula1=yes_no_values[1])
                self._ws1.add_data_validation(data_vals)
        elif language == "EN":
            for column in columns:
                data_vals = DataValidation(sqref=column, type="list", formula1=yes_no_values[2])
                self._ws1.add_data_validation(data_vals)

    def fill_in_number_values(self):
        cz_number_column_letters = []
        sk_number_column_letters = []
        en_number_column_letters = []

        for i, types in enumerate(self._second_row_type):
            if types == "Numerická hodnota":
                if self._third_row_language[i] == "SK":
                    sk_number_column_letters.append(get_column_letter(i + 1))
                elif self._third_row_language[i] == "EN":
                    en_number_column_letters.append(get_column_letter(i + 1))
                elif self._third_row_language[i] == "CZ":
                    cz_number_column_letters.append(get_column_letter(i + 1))

            if self._third_row_language[i] == "SK":
                for ind, cz_column in enumerate(cz_number_column_letters):
                    for j in range(5, self._settings[3] + 1):
                        self._ws1[
                            f"{sk_number_column_letters[ind]}{j}"] = f'=if({cz_column}{j}="","",{cz_column}{j})'
            elif self._third_row_language[i] == "EN":
                for ind, cz_column in enumerate(cz_number_column_letters):
                    for j in range(5, self._settings[3] + 1):
                        self._ws1[
                            f"{en_number_column_letters[ind]}{j}"] = f'=if({cz_column}{j}="","",SUBSTITUTE({cz_column}{j},",","."))'

    def fill_in_yes_no_values(self):
        cz_column_letters = []
        sk_column_letters = []
        en_column_letters = []

        for i, types in enumerate(self._second_row_type):
            if types == "Ano/Ne":
                if self._third_row_language[i] == "SK":
                    sk_column_letters.append(get_column_letter(i + 1))
                elif self._third_row_language[i] == "EN":
                    en_column_letters.append(get_column_letter(i + 1))
                elif self._third_row_language[i] == "CZ":
                    cz_column_letters.append(get_column_letter(i + 1))

            if self._third_row_language[i] == "SK":
                for ind, cz_column in enumerate(cz_column_letters):
                    for j in range(5, self._settings[3] + 1):
                        self._ws1[
                            f"{sk_column_letters[ind]}{j}"] = f'=if({cz_column}{j}="","",if({cz_column}{j}="Ano","Áno",if({cz_column}{j}="Ne","Nie")))'
            elif self._third_row_language[i] == "EN":
                for ind, cz_column in enumerate(cz_column_letters):
                    for j in range(5, self._settings[3] + 1):
                        self._ws1[
                            f"{en_column_letters[ind]}{j}"] = f'=if({cz_column}{j}="","",if({cz_column}{j}="Ano","Yes",if({cz_column}{j}="Ne","No")))'

    def format_cells(self):
        for i in range(5, self._settings[
                              3] + 1):  # this adds value from the choice variable to lines 5 to value from settings
            self._ws1.append(self._fifth_row_menu_data)

        MIN_WIDTH = 10
        for i, column_cells in enumerate(self._ws1.columns, start=1):
            width = (
                length
                if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                                  for cell in column_cells)) >= MIN_WIDTH
                else MIN_WIDTH
            )
            self._ws1.column_dimensions[get_column_letter(i)].width = width + 1

        for cell in self._ws1[4:4]:
            cell.font = Font(bold=True)

        for row in range(1, 3):
            for cell in self._ws1[row]:
                cell.alignment = openpyxl.styles.Alignment(horizontal="left")

    def write_file(self):
        self._ws1.title = self._choice
        self._ws1.append(self._first_row_unit)
        self._ws1.append(self._second_row_type)
        self._ws1.append(self._third_row_language)
        self._ws1.append(self._fourth_row_parameter)
        self.format_cells()
        self._ws1.freeze_panes = "C5"
        for language in Pargen.LANGUAGE_NAMES:
            self.create_yes_no_dropdowns(language)
        self.create_text_parameters_dropdowns()
        self.fill_in_yes_no_values()
        self.fill_in_number_values()
        output_file = f"{self._choice}.xlsx"
        if os.path.isfile(output_file) is True:
            question = input(str(f'Soubor "{output_file}"{Pargen.FILE_OVERWRITE_TEXT}'))
            if question == 'a' or question == 'A':
                self._wb_output.save(filename=output_file)
        else:
            self._wb_output.save(filename=output_file)


if __name__ == "__main__":
    er = Pargen()
    er.prepare_data()
    er.write_file()
